import openpyxl

# step 1 create the workbook
book=openpyxl.Workbook()

#step 2 set the active sheet
sheet = book.active

#step 3 add the data in cells
cell = sheet.cell(row=1,column=1)
cell1 =  sheet.cell(row=1,column=2)
cell.value="Ganga"
cell1.value="s"


#add entire table in excel
developers = [["name","job"],["ganga","SE"],["swati","Trainer"],["sneha","lead"]]
maxRow = len(developers)
maxCol = len(developers[0])

for i in range(0,maxRow):
    for j in range(0,maxCol):
        cell=sheet.cell(row=i+1,column=j+1)
        cell.value=developers[i][j]


#step4 save file
book.save("/home/redintegro/Documents/dev.xlsx")