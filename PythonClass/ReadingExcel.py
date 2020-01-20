import openpyxl

#step 1
book=openpyxl.load_workbook("/home/redintegro/Documents/demo.xlsx")

#step 2 : set active sheet
sheet= book.active

#read value from cell (single cell)
# cell=sheet['A1']   
# cell=sheet.cell(row=5,column=4)             
# print(cell.value)


#reading entire row
# maxCol = sheet.max_column
# print(maxCol)
# for i in range(1,maxCol+1):
#     cell=sheet.cell(row=12,column=i)
#     print(cell.value)

#reading entire table
maxCol = sheet.max_column
maxRow = sheet.max_row
for i in range(1,maxRow):
    for j in range(1,maxCol):
        cell=sheet.cell(row=i,column=j)
        print(cell.value)
